#!/usr/bin/env python3
"""
Check all sheets in Excel file for South Texas Onion Committee mentions
"""
import openpyxl
from datetime import datetime

excel_path = '/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx'
wb = openpyxl.load_workbook(excel_path)

print("="*80)
print("SEARCHING ALL SHEETS FOR ONION/SOUTH TEXAS MENTIONS")
print("="*80)
print()

print("Available sheets:", wb.sheetnames)
print()

keywords = ['onion', 'texas', 'south texas', '1015', 'tx1015']

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"CHECKING SHEET: {sheet_name}")
    print(f"{'='*80}")

    sheet = wb[sheet_name]
    all_rows = list(sheet.iter_rows(values_only=True))

    print(f"Total rows: {len(all_rows)}")

    # Search for keywords
    matches_found = []

    for row_idx, row in enumerate(all_rows, start=1):
        for col_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                cell_lower = cell.lower()
                for keyword in keywords:
                    if keyword in cell_lower:
                        matches_found.append({
                            'row': row_idx,
                            'col': col_idx,
                            'keyword': keyword,
                            'cell': cell,
                            'full_row': row
                        })
                        break  # Only count once per cell

    if matches_found:
        print(f"\nFound {len(matches_found)} cells with keywords")
        print("\nMatches:")
        for match in matches_found[:20]:  # Show first 20
            print(f"\n  Row {match['row']}, Col {match['col']}: Keyword '{match['keyword']}'")
            print(f"    Cell content: {str(match['cell'])[:100]}")

            # Check if this row has a date
            date_val = None
            for cell in match['full_row'][:5]:
                if isinstance(cell, datetime):
                    date_val = cell
                    break
            if date_val:
                print(f"    Date in row: {date_val.date()}")
    else:
        print("  No keyword matches found")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
