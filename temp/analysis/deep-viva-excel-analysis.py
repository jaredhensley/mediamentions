#!/usr/bin/env python3
"""
Deep analysis of Viva tab - check all dates and content
"""
import openpyxl
from datetime import datetime
import json

excel_path = '/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx'
wb = openpyxl.load_workbook(excel_path)

if 'Viva' not in wb.sheetnames:
    print("No Viva sheet found")
    exit(1)

sheet = wb['Viva']

print("="*80)
print("COMPLETE VIVA SHEET ANALYSIS")
print("="*80)
print()

# Get all rows
all_rows = list(sheet.iter_rows(values_only=True))

print(f"Total rows in sheet: {len(all_rows)}")
print()

# Print all rows with row numbers
print("ALL ROWS (RAW DATA):")
print("-"*80)
for i, row in enumerate(all_rows, start=1):
    # Only print rows that have at least one non-None value
    if any(cell is not None for cell in row):
        print(f"\nRow {i}:")
        for j, cell in enumerate(row[:15]):  # First 15 columns
            if cell is not None:
                cell_type = type(cell).__name__
                if isinstance(cell, datetime):
                    print(f"  Col {j} ({cell_type}): {cell} [{cell.date()}]")
                else:
                    cell_str = str(cell)[:200]  # Limit to 200 chars
                    print(f"  Col {j} ({cell_type}): {cell_str}")

print("\n" + "="*80)
print("DATE ANALYSIS")
print("="*80)

# Check all cells for dates
dates_found = []
for row_idx, row in enumerate(all_rows, start=1):
    for col_idx, cell in enumerate(row):
        if isinstance(cell, datetime):
            dates_found.append({
                'row': row_idx,
                'col': col_idx,
                'date': cell,
                'year': cell.year
            })

print(f"\nTotal date cells found: {len(dates_found)}")

if dates_found:
    print("\nDates by year:")
    years = {}
    for d in dates_found:
        year = d['year']
        if year not in years:
            years[year] = []
        years[year].append(d)

    for year in sorted(years.keys()):
        print(f"  {year}: {len(years[year])} dates")
        for d in years[year]:
            print(f"    Row {d['row']}, Col {d['col']}: {d['date'].date()}")

print("\n" + "="*80)
print("CHECKING FOR 2025 DATA")
print("="*80)

dates_2025 = [d for d in dates_found if d['year'] == 2025]
print(f"\nDates in 2025: {len(dates_2025)}")

if dates_2025:
    print("\nAll 2025 dates:")
    for d in sorted(dates_2025, key=lambda x: x['date']):
        print(f"  Row {d['row']}, Col {d['col']}: {d['date'].date()}")
        # Print the full row for context
        row_data = all_rows[d['row']-1]
        print(f"    Full row: {row_data[:10]}")

# Check for any text mentions of dates in 2025
print("\n" + "="*80)
print("TEXT SEARCH FOR 2025 PATTERNS")
print("="*80)

for row_idx, row in enumerate(all_rows, start=1):
    for col_idx, cell in enumerate(row):
        if isinstance(cell, str) and '2025' in cell:
            print(f"\nRow {row_idx}, Col {col_idx}: Found '2025' in text")
            print(f"  Content: {cell}")
