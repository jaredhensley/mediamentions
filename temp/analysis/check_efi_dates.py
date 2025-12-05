#!/usr/bin/env python3
"""
Check what date range is actually in the EFI Excel file
"""

import openpyxl
from datetime import datetime

EXCEL_FILE = "/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx"

def check_dates():
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Find EFI sheet
    efi_sheet = None
    for sheet_name in wb.sheetnames:
        if 'EFI' in sheet_name.upper():
            efi_sheet = wb[sheet_name]
            break

    if not efi_sheet:
        print("No EFI sheet found")
        return

    print("Checking dates in EFI sheet...\n")

    dates = []
    for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=5, values_only=True), start=5):
        if not any(row):
            continue

        date_val = row[0]  # Col 0 is Date
        if isinstance(date_val, datetime):
            dates.append(date_val)
            if len(dates) <= 10 or len(dates) % 20 == 0:
                print(f"Row {row_idx}: {date_val.strftime('%Y-%m-%d')}")

    if dates:
        print(f"\n{'=' * 60}")
        print(f"Total dates found: {len(dates)}")
        print(f"Earliest date: {min(dates).strftime('%Y-%m-%d')}")
        print(f"Latest date: {max(dates).strftime('%Y-%m-%d')}")

        # Count by year
        from collections import defaultdict
        year_counts = defaultdict(int)
        for d in dates:
            year_counts[d.year] += 1

        print(f"\nBy year:")
        for year in sorted(year_counts.keys()):
            print(f"  {year}: {year_counts[year]} mentions")

        # Check for June 7 - Dec 4 2024 window
        start_2024 = datetime(2024, 6, 7)
        end_2024 = datetime(2024, 12, 4)
        count_2024 = sum(1 for d in dates if start_2024 <= d <= end_2024)
        print(f"\nMentions in 2024-06-07 to 2024-12-04: {count_2024}")

        # Check for the 180 days before the LATEST date
        if dates:
            latest = max(dates)
            from datetime import timedelta
            window_start = latest - timedelta(days=180)
            count_window = sum(1 for d in dates if window_start <= d <= latest)
            print(f"\nMentions in 180-day window before {latest.strftime('%Y-%m-%d')}:")
            print(f"  {window_start.strftime('%Y-%m-%d')} to {latest.strftime('%Y-%m-%d')}: {count_window} mentions")

if __name__ == '__main__':
    check_dates()
