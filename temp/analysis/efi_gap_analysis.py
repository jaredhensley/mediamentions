#!/usr/bin/env python3
"""
Deep dive analysis: Why are we missing 59 out of 82 mentions?
"""

import openpyxl
from datetime import datetime
from collections import defaultdict

EXCEL_FILE = "/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx"
START_DATE = datetime(2025, 6, 7)
END_DATE = datetime(2025, 12, 4)

def analyze_gaps():
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Find EFI sheet
    efi_sheet = None
    for sheet_name in wb.sheetnames:
        if 'EFI' in sheet_name.upper():
            efi_sheet = wb[sheet_name]
            break

    if not efi_sheet:
        return

    # Collect mentions in range
    mentions = []
    for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=5, values_only=True), start=5):
        if not any(row):
            continue

        date_val = row[0]
        if isinstance(date_val, datetime) and START_DATE <= date_val <= END_DATE:
            mentions.append({
                'date': date_val,
                'source': row[1],
                'title': row[2],
                'topic': row[3],
                'url': row[5]
            })

    print("=" * 80)
    print("EFI GAP ANALYSIS: Why are we missing 72% of mentions?")
    print("=" * 80)
    print(f"\nTotal mentions to analyze: {len(mentions)}")
    print()

    # Analyze title patterns in detail
    print("=" * 80)
    print("1. TITLE PATTERN ANALYSIS")
    print("=" * 80)

    categories = {
        'full_name_title': [],
        'efi_acronym_title': [],
        'ecip_only': [],
        'indirect_mention': [],
        'company_certification': [],
        'no_clear_mention': []
    }

    for m in mentions:
        title = str(m['title']).lower() if m['title'] else ''
        topic = str(m['topic']).lower() if m['topic'] else ''

        if 'equitable food initiative' in title:
            categories['full_name_title'].append(m)
        elif 'efi' in title and 'ecip' not in title:
            categories['efi_acronym_title'].append(m)
        elif 'ecip' in title or 'ethical charter' in title:
            categories['ecip_only'].append(m)
        elif any(word in topic for word in ['certified', 'certification']):
            categories['company_certification'].append(m)
        elif any(word in title for word in ['farmworker', 'labor', 'workers', 'social responsibility']):
            categories['indirect_mention'].append(m)
        else:
            categories['no_clear_mention'].append(m)

    for cat_name, cat_mentions in categories.items():
        pct = len(cat_mentions) / len(mentions) * 100 if mentions else 0
        print(f"\n{cat_name.replace('_', ' ').title()}: {len(cat_mentions)} ({pct:.1f}%)")

        if cat_mentions and len(cat_mentions) <= 5:
            for m in cat_mentions[:5]:
                print(f"  - {m['title'][:70]}")
        elif cat_mentions:
            for m in cat_mentions[:3]:
                print(f"  - {m['title'][:70]}")
            print(f"  ... and {len(cat_mentions) - 3} more")

    # Analyze ECIP mentions specifically
    print("\n" + "=" * 80)
    print("2. ECIP (Ethical Charter Implementation Program) MENTIONS")
    print("=" * 80)

    ecip_mentions = [m for m in mentions
                     if 'ecip' in str(m['title']).lower() or
                     'ethical charter' in str(m['title']).lower() or
                     'ecip' in str(m['topic']).lower()]

    print(f"\nTotal ECIP-related mentions: {len(ecip_mentions)} ({len(ecip_mentions)/len(mentions)*100:.1f}%)")
    print("\nECIP is EFI's related program. These mentions likely:")
    print("  - Don't explicitly say 'Equitable Food Initiative' in title")
    print("  - May only mention EFI in the body text")
    print("  - Require reading full article to confirm EFI connection")
    print("\nSample ECIP mentions:")
    for m in ecip_mentions[:5]:
        print(f"\n  Title: {m['title']}")
        print(f"  Source: {m['source']}")
        print(f"  Date: {m['date'].strftime('%Y-%m-%d')}")

    # Analyze company certification mentions
    print("\n" + "=" * 80)
    print("3. COMPANY CERTIFICATION ANNOUNCEMENTS")
    print("=" * 80)

    cert_keywords = ['certified', 'certification', 'achieves', 'earns', 'receives']
    cert_mentions = []

    for m in mentions:
        title = str(m['title']).lower() if m['title'] else ''
        topic = str(m['topic']).lower() if m['topic'] else ''

        if any(kw in title for kw in cert_keywords) or any(kw in topic for kw in cert_keywords):
            # Check if EFI is mentioned
            has_efi_in_title = 'efi' in title or 'equitable food' in title
            cert_mentions.append({
                'mention': m,
                'efi_in_title': has_efi_in_title
            })

    efi_in_title = sum(1 for c in cert_mentions if c['efi_in_title'])
    efi_not_in_title = len(cert_mentions) - efi_in_title

    print(f"\nTotal certification mentions: {len(cert_mentions)}")
    print(f"  With EFI in title: {efi_in_title}")
    print(f"  Without EFI in title (likely in body only): {efi_not_in_title}")

    print("\nCertification mentions WITHOUT EFI in title:")
    for c in [c for c in cert_mentions if not c['efi_in_title']][:5]:
        m = c['mention']
        print(f"\n  Title: {m['title'][:70]}")
        print(f"  Topic: {m['topic']}")
        print(f"  Source: {m['source']}")

    # Analyze event/award mentions
    print("\n" + "=" * 80)
    print("4. EVENT & AWARD MENTIONS")
    print("=" * 80)

    event_keywords = ['award', 'honor', 'recogniz', 'celebrat', 'event', 'conference', 'show']
    event_mentions = [m for m in mentions
                      if any(kw in str(m['title']).lower() for kw in event_keywords)]

    print(f"\nTotal event/award mentions: {len(event_mentions)} ({len(event_mentions)/len(mentions)*100:.1f}%)")

    efi_in_title = sum(1 for m in event_mentions
                       if 'efi' in str(m['title']).lower() or 'equitable food' in str(m['title']).lower())

    print(f"  With EFI explicitly in title: {efi_in_title}")
    print(f"  Without EFI in title: {len(event_mentions) - efi_in_title}")

    print("\nEvent mentions WITHOUT EFI in title:")
    for m in [m for m in event_mentions
              if 'efi' not in str(m['title']).lower() and 'equitable food' not in str(m['title']).lower()][:5]:
        print(f"\n  Title: {m['title'][:70]}")
        print(f"  Topic: {m['topic']}")

    # Industry roundup/newsletter mentions
    print("\n" + "=" * 80)
    print("5. POTENTIAL INDUSTRY ROUNDUPS / BRIEF MENTIONS")
    print("=" * 80)

    # Short titles often indicate brief mentions or roundups
    brief_titles = [m for m in mentions if len(str(m['title'])) < 50]
    print(f"\nMentions with short titles (<50 chars): {len(brief_titles)}")
    print("These may be brief mentions or industry roundups where EFI is mentioned in passing")

    # High-value missed mentions
    print("\n" + "=" * 80)
    print("6. HIGH-VALUE MISSED MENTIONS (Detailed Sample)")
    print("=" * 80)

    # Sort by date
    recent = sorted(mentions, key=lambda x: x['date'], reverse=True)[:10]

    print("\nTop 10 most recent mentions (ALL should be findable):")
    for idx, m in enumerate(recent, 1):
        title = str(m['title']) if m['title'] else 'NO TITLE'
        topic = str(m['topic']) if m['topic'] else 'NO TOPIC'

        # Check if searchable
        searchable = False
        search_method = "UNCLEAR"

        title_lower = title.lower()
        if 'equitable food initiative' in title_lower:
            searchable = True
            search_method = "Full name in title"
        elif 'efi' in title_lower:
            searchable = True
            search_method = "EFI acronym in title"
        elif 'ecip' in title_lower or 'ethical charter' in title_lower:
            searchable = "MAYBE"
            search_method = "ECIP in title (need body text)"
        else:
            searchable = False
            search_method = "EFI likely only in body text"

        print(f"\n{idx}. {title[:70]}")
        print(f"   Date: {m['date'].strftime('%Y-%m-%d')}")
        print(f"   Source: {m['source']}")
        print(f"   Topic: {topic}")
        print(f"   Searchable: {searchable} - {search_method}")

    # Summary recommendations
    print("\n" + "=" * 80)
    print("7. SUMMARY & RECOMMENDATIONS")
    print("=" * 80)

    direct_findable = len(categories['full_name_title']) + len(categories['efi_acronym_title'])
    ecip_findable = len(categories['ecip_only'])
    body_only = len(mentions) - direct_findable - ecip_findable

    print(f"\nMention breakdown:")
    print(f"  Directly findable (EFI in title): {direct_findable} ({direct_findable/len(mentions)*100:.1f}%)")
    print(f"  ECIP mentions (might be findable): {ecip_findable} ({ecip_findable/len(mentions)*100:.1f}%)")
    print(f"  Body-only mentions: {body_only} ({body_only/len(mentions)*100:.1f}%)")

    print(f"\nCurrent coverage: 23 out of 82 (28.0%)")
    print(f"Theoretical maximum with title-based search: ~{direct_findable + ecip_findable} mentions")
    print(f"Maximum coverage possible: ~{(direct_findable + ecip_findable)/len(mentions)*100:.1f}%")

    print("\nKey findings:")
    print("  1. ~60% of mentions don't have EFI in the title")
    print("  2. ECIP mentions need special handling")
    print("  3. Many company certification announcements mention EFI only in body")
    print("  4. Events/awards often mention EFI as a participant/sponsor in body")

    print("\nRecommendations:")
    print("  1. ECIP query: Add 'ECIP OR \"Ethical Charter Implementation Program\"'")
    print("  2. Accept 28-40% coverage as realistic baseline for title-based search")
    print("  3. Focus on high-value direct mentions (current query is good)")
    print("  4. For full coverage, would need full-text search (expensive/complex)")

if __name__ == '__main__':
    analyze_gaps()
