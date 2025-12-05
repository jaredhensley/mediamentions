#!/usr/bin/env python3
"""
Inspect the EFI Excel file structure
"""

import openpyxl
from datetime import datetime

EXCEL_FILE = "/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx"

def inspect_excel():
    """Inspect the Excel file structure"""

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Find EFI sheet
    efi_sheet = None
    for sheet_name in wb.sheetnames:
        if 'EFI' in sheet_name.upper():
            efi_sheet = wb[sheet_name]
            print(f"Found sheet: '{sheet_name}'")
            break

    if not efi_sheet:
        print(f"No EFI sheet found. Available sheets: {wb.sheetnames}")
        return

    print(f"\n{'=' * 80}")
    print("First 10 rows of the sheet:")
    print("=" * 80)

    for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        print(f"\nRow {row_idx}:")
        non_empty = [(idx, val) for idx, val in enumerate(row) if val is not None]
        if non_empty:
            for idx, val in non_empty:
                val_str = str(val)[:100]
                print(f"  Col {idx}: {val_str}")
        else:
            print("  (empty row)")

    print(f"\n{'=' * 80}")
    print("Scanning for header patterns...")
    print("=" * 80)

    for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_str = ' | '.join([str(v) if v else '' for v in row[:10]])
        if any(keyword in row_str.lower() for keyword in ['date', 'title', 'source', 'headline', 'url', 'publication']):
            print(f"\nRow {row_idx} (possible header): {row_str}")

    print(f"\n{'=' * 80}")
    print("Checking data pattern starting from different rows...")
    print("=" * 80)

    for start_row in [2, 3, 4, 5]:
        print(f"\nStarting from row {start_row}:")
        count = 0
        for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=start_row, max_row=start_row+5, values_only=True), start=start_row):
            if any(row):
                count += 1
                # Show first 5 non-empty values
                non_empty = [str(v)[:50] for v in row if v is not None][:5]
                print(f"  Row {row_idx}: {' | '.join(non_empty)}")
        print(f"  Non-empty rows found: {count}")

if __name__ == '__main__':
    inspect_excel()
