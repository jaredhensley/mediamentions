#!/usr/bin/env python3
"""
Analyze Viva client mentions from Excel file
"""
import openpyxl
from datetime import datetime, timedelta
import json

# Load the Excel file
excel_path = '/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx'
wb = openpyxl.load_workbook(excel_path)

# Check if "Viva" sheet exists
print("Available sheets:", wb.sheetnames)
print()

if 'Viva' in wb.sheetnames:
    sheet = wb['Viva']

    # Get headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)

    print("Column headers:")
    for i, h in enumerate(headers):
        print(f"{i}: {h}")
    print()

    # Define date range: June 7 - December 4, 2025
    start_date = datetime(2025, 6, 7)
    end_date = datetime(2025, 12, 4)

    print(f"Filtering mentions from {start_date.date()} to {end_date.date()}")
    print()

    # Collect all mentions
    all_mentions = []
    filtered_mentions = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        values = [cell.value for cell in row]

        # Try to find date column (usually first few columns)
        date_val = None
        for col_idx in range(min(5, len(values))):
            val = values[col_idx]
            if val:
                # Try parsing as date
                if isinstance(val, datetime):
                    date_val = val
                    break
                elif isinstance(val, str):
                    # Try parsing string dates
                    for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%Y']:
                        try:
                            date_val = datetime.strptime(val.strip(), fmt)
                            break
                        except:
                            continue
                    if date_val:
                        break

        # Create mention record
        mention = {
            'row': row_idx,
            'date': date_val,
            'values': values[:10] if len(values) > 10 else values  # First 10 columns
        }

        all_mentions.append(mention)

        # Filter by date range
        if date_val and start_date <= date_val <= end_date:
            filtered_mentions.append(mention)

    print(f"Total rows in Viva sheet: {len(all_mentions)}")
    print(f"Mentions in 180-day window (June 7 - Dec 4, 2025): {len(filtered_mentions)}")
    print()

    # Show sample of filtered mentions
    print("Sample of filtered mentions (first 5):")
    for i, m in enumerate(filtered_mentions[:5]):
        print(f"\nRow {m['row']}: Date={m['date'].date() if m['date'] else 'N/A'}")
        for j, v in enumerate(m['values']):
            if v and j < len(headers):
                print(f"  {headers[j]}: {str(v)[:100]}")

    # Look for client identification in headers/data
    print("\n" + "="*80)
    print("CHECKING CLIENT IDENTIFICATION")
    print("="*80)

    # Check first few rows for client name references
    print("\nFirst 3 rows (full data):")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
        print(f"\nRow {row_idx}:")
        for i, val in enumerate(row[:15]):  # First 15 columns
            if val and i < len(headers):
                print(f"  {headers[i]}: {val}")

    # Save detailed analysis
    output = {
        'total_mentions': len(all_mentions),
        'filtered_mentions_count': len(filtered_mentions),
        'date_range': {
            'start': start_date.isoformat(),
            'end': end_date.isoformat()
        },
        'headers': headers,
        'sample_mentions': []
    }

    # Add detailed sample with all fields
    for m in filtered_mentions[:20]:
        sample = {
            'row': m['row'],
            'date': m['date'].isoformat() if m['date'] else None,
        }
        for i, val in enumerate(m['values']):
            if i < len(headers) and headers[i]:
                sample[headers[i]] = str(val) if val else None
        output['sample_mentions'].append(sample)

    with open('/Users/jaredhensley/Code/mediamentions/viva-analysis.json', 'w') as f:
        json.dump(output, f, indent=2)

    print("\n\nDetailed analysis saved to: viva-analysis.json")
    print(f"\nMANUAL TRACKING COUNT (180-day window): {len(filtered_mentions)}")

else:
    print("ERROR: 'Viva' sheet not found in workbook")
    print(f"Available sheets: {wb.sheetnames}")
