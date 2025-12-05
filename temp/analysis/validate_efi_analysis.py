#!/usr/bin/env python3
"""
EFI Analysis Validation Script
Analyzes manual tracking data vs automated search results
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import re

# File paths
EXCEL_FILE = "/Users/jaredhensley/Downloads/Media Mentions - All Clients (1).xlsx"
# The user mentioned June 7 - Dec 4, 2025 (180-day window)
START_DATE = datetime(2025, 6, 7)
END_DATE = datetime(2025, 12, 4)

def parse_date(date_value):
    """Parse date from Excel cell"""
    if isinstance(date_value, datetime):
        return date_value
    if isinstance(date_value, str):
        # Try various date formats
        formats = [
            '%m/%d/%Y',
            '%Y-%m-%d',
            '%m/%d/%y',
            '%d-%b-%y',
            '%b %d, %Y',
            '%B %d, %Y'
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_value, fmt)
            except ValueError:
                continue
    return None

def analyze_efi_sheet():
    """Analyze the EFI tab from the Excel file"""

    print("=" * 80)
    print("EFI ANALYSIS VALIDATION REPORT")
    print("=" * 80)
    print(f"\nAnalysis Period: {START_DATE.strftime('%B %d, %Y')} - {END_DATE.strftime('%B %d, %Y')}")
    print(f"Duration: {(END_DATE - START_DATE).days} days\n")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Find EFI sheet
    efi_sheet = None
    for sheet_name in wb.sheetnames:
        if 'EFI' in sheet_name.upper():
            efi_sheet = wb[sheet_name]
            print(f"Found sheet: '{sheet_name}'")
            break

    if not efi_sheet:
        print("ERROR: Could not find EFI sheet!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    # Parse headers from row 2
    headers = []
    for cell in efi_sheet[2]:  # Row 2 has headers
        headers.append(cell.value)

    print(f"\nColumns found: {headers}\n")

    # Column mapping based on inspection
    # Col 0: Date, Col 1: Publication Name, Col 2: Title, Col 3: Topic, Col 4: Additional Mentions, Col 5: Link
    date_col = 0
    source_col = 1
    title_col = 2
    type_col = 3
    url_col = 5

    print(f"Column mapping:")
    print(f"  Date column: {date_col} ({headers[date_col] if date_col is not None else 'NOT FOUND'})")
    print(f"  Title column: {title_col} ({headers[title_col] if title_col is not None else 'NOT FOUND'})")
    print(f"  Source column: {source_col} ({headers[source_col] if source_col is not None else 'NOT FOUND'})")
    print(f"  URL column: {url_col} ({headers[url_col] if url_col is not None else 'NOT FOUND'})")
    print(f"  Type column: {type_col} ({headers[type_col] if type_col is not None else 'NOT FOUND'})")

    # Collect mentions
    all_mentions = []
    mentions_in_range = []
    monthly_counts = defaultdict(int)
    source_counts = defaultdict(int)
    type_counts = defaultdict(int)

    # Skip rows 1-4 (title, headers, example, template)
    for row_idx, row in enumerate(efi_sheet.iter_rows(min_row=5, values_only=True), start=5):
        if not any(row):  # Skip empty rows
            continue

        date_val = row[date_col] if date_col is not None else None
        title_val = row[title_col] if title_col is not None else None
        source_val = row[source_col] if source_col is not None else None
        url_val = row[url_col] if url_col is not None else None
        type_val = row[type_col] if type_col is not None else None

        # Parse date
        mention_date = parse_date(date_val)

        mention = {
            'row': row_idx,
            'date': mention_date,
            'date_str': str(date_val),
            'title': title_val,
            'source': source_val,
            'url': url_val,
            'type': type_val
        }

        all_mentions.append(mention)

        # Check if in range
        if mention_date and START_DATE <= mention_date <= END_DATE:
            mentions_in_range.append(mention)

            # Count by month
            month_key = mention_date.strftime('%Y-%m')
            monthly_counts[month_key] += 1

            # Count by source
            if source_val:
                source_counts[str(source_val).strip()] += 1

            # Count by type
            if type_val:
                type_counts[str(type_val).strip()] += 1

    print(f"\n{'=' * 80}")
    print("1. MANUAL TRACKING VERIFICATION")
    print("=" * 80)
    print(f"\nTotal mentions in Excel: {len(all_mentions)}")
    print(f"Mentions in date range ({START_DATE.strftime('%m/%d/%Y')} - {END_DATE.strftime('%m/%d/%Y')}): {len(mentions_in_range)}")

    # Monthly breakdown
    print(f"\n{'=' * 80}")
    print("MONTHLY BREAKDOWN")
    print("=" * 80)
    for month in sorted(monthly_counts.keys()):
        count = monthly_counts[month]
        month_date = datetime.strptime(month, '%Y-%m')
        print(f"{month_date.strftime('%B %Y'):20s}: {count:3d} mentions")

    # Source breakdown
    print(f"\n{'=' * 80}")
    print("TOP SOURCES")
    print("=" * 80)
    sorted_sources = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)
    for source, count in sorted_sources[:15]:
        print(f"{source[:50]:50s}: {count:3d} mentions")

    # Type breakdown
    if type_counts:
        print(f"\n{'=' * 80}")
        print("MENTION TYPES")
        print("=" * 80)
        sorted_types = sorted(type_counts.items(), key=lambda x: x[1], reverse=True)
        for mtype, count in sorted_types:
            print(f"{str(mtype):30s}: {count:3d} mentions")

    # High-value missed mentions
    print(f"\n{'=' * 80}")
    print("2. TOP 20 MISSED MENTIONS (MANUAL TRACKING)")
    print("=" * 80)
    print("\nSorted by date (most recent first):\n")

    sorted_mentions = sorted(mentions_in_range, key=lambda x: x['date'], reverse=True)

    for idx, mention in enumerate(sorted_mentions[:20], 1):
        date_str = mention['date'].strftime('%m/%d/%Y') if mention['date'] else mention['date_str']
        title = str(mention['title'])[:80] if mention['title'] else 'NO TITLE'
        source = str(mention['source'])[:40] if mention['source'] else 'NO SOURCE'

        print(f"{idx:2d}. Date: {date_str}")
        print(f"    Title: {title}")
        print(f"    Source: {source}")
        if mention['url']:
            print(f"    URL: {str(mention['url'])[:100]}")
        print()

    # Pattern analysis
    print(f"\n{'=' * 80}")
    print("3. PATTERN ANALYSIS")
    print("=" * 80)

    # Analyze titles for patterns
    efi_full_name = 0
    efi_acronym = 0
    other_patterns = 0

    for mention in mentions_in_range:
        title_lower = str(mention['title']).lower() if mention['title'] else ''

        if 'equitable food initiative' in title_lower:
            efi_full_name += 1
        elif 'efi' in title_lower:
            efi_acronym += 1
        else:
            other_patterns += 1

    if len(mentions_in_range) > 0:
        print(f"\nTitle Pattern Analysis:")
        print(f"  Contains 'Equitable Food Initiative': {efi_full_name} ({efi_full_name/len(mentions_in_range)*100:.1f}%)")
        print(f"  Contains 'EFI' only: {efi_acronym} ({efi_acronym/len(mentions_in_range)*100:.1f}%)")
        print(f"  Neither (may be in body only): {other_patterns} ({other_patterns/len(mentions_in_range)*100:.1f}%)")
    else:
        print(f"\nNo mentions found in date range - cannot perform pattern analysis")

    # Analyze URLs for patterns
    url_patterns = defaultdict(int)
    for mention in mentions_in_range:
        if mention['url']:
            url_str = str(mention['url'])
            # Extract domain
            import re
            domain_match = re.search(r'https?://(?:www\.)?([^/]+)', url_str)
            if domain_match:
                domain = domain_match.group(1)
                url_patterns[domain] += 1

    print(f"\nTop URL Domains:")
    sorted_domains = sorted(url_patterns.items(), key=lambda x: x[1], reverse=True)
    for domain, count in sorted_domains[:10]:
        print(f"  {domain:40s}: {count:3d} mentions")

    # Coverage analysis
    print(f"\n{'=' * 80}")
    print("4. COVERAGE ANALYSIS")
    print("=" * 80)

    automated_count = 23
    manual_count = len(mentions_in_range)
    missed_count = manual_count - automated_count
    coverage_pct = (automated_count / manual_count * 100) if manual_count > 0 else 0

    print(f"\nAutomated search results: {automated_count}")
    print(f"Manual tracking mentions: {manual_count}")
    print(f"Missed mentions: {missed_count}")
    print(f"Coverage rate: {coverage_pct:.1f}%")
    print(f"\nImprovement: From 1 result (1.2%) to {automated_count} results ({coverage_pct:.1f}%)")

    # Sample of specific missed mentions by category
    print(f"\n{'=' * 80}")
    print("5. SAMPLE MISSED MENTIONS BY SOURCE TYPE")
    print("=" * 80)

    # Group by source patterns
    news_sources = []
    trade_sources = []
    other_sources = []

    trade_keywords = ['produce', 'fresh', 'grower', 'packer', 'food safety', 'agriculture',
                      'agri', 'farm', 'harvest', 'retail']

    for mention in sorted_mentions[:40]:  # Check top 40
        source = str(mention['source']).lower() if mention['source'] else ''

        is_trade = any(keyword in source for keyword in trade_keywords)

        if is_trade:
            trade_sources.append(mention)
        elif mention['source']:
            news_sources.append(mention)
        else:
            other_sources.append(mention)

    print(f"\nTrade/Industry Publications (sample of {min(5, len(trade_sources))}):")
    for idx, mention in enumerate(trade_sources[:5], 1):
        date_str = mention['date'].strftime('%m/%d/%Y') if mention['date'] else mention['date_str']
        print(f"  {idx}. {mention['source']}")
        print(f"     Title: {str(mention['title'])[:70]}")
        print(f"     Date: {date_str}")
        print()

    print(f"News/General Publications (sample of {min(5, len(news_sources))}):")
    for idx, mention in enumerate(news_sources[:5], 1):
        date_str = mention['date'].strftime('%m/%d/%Y') if mention['date'] else mention['date_str']
        print(f"  {idx}. {mention['source']}")
        print(f"     Title: {str(mention['title'])[:70]}")
        print(f"     Date: {date_str}")
        print()

    return {
        'total_in_range': len(mentions_in_range),
        'monthly_counts': monthly_counts,
        'source_counts': source_counts,
        'mentions': mentions_in_range
    }

if __name__ == '__main__':
    analyze_efi_sheet()
